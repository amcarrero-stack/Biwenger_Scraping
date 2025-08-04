# bloque_1_inicializacion.py

from utils import ensure_directories_exist, get_excel_path, log_message
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
from config import SALDO_INICIAL, NOMBRE_MI_EQUIPO

def inicializar_bbdd_y_directorios():
    ensure_directories_exist()
    excel_path = get_excel_path()

    try:
        wb = openpyxl.load_workbook(excel_path)
        log_message(f"Archivo Excel existente cargado: {excel_path}")
    except FileNotFoundError:
        log_message("No existe archivo Excel. Creando nuevo archivo y hoja Principal e Historial...")
        wb = Workbook()

        # Crear hoja Principal
        ws_principal = wb.active
        ws_principal.title = "Principal"
        # Encabezados
        ws_principal.append(["Nombre", "Nº jugadores", "Saldo", "fecha/hora"])
        # Aquí deberías añadir los usuarios excepto el tuyo, con saldo inicial de 20M
        usuarios = ["Al-khelaifi", "Juanjo", "Yyoquese", "Bellingham5", "Jaime Palomino Cano", "Mast-antonio", "david"]
        for u in usuarios:
            if u != NOMBRE_MI_EQUIPO:
                ws_principal.append([u, 0, SALDO_INICIAL, ""])  # fecha/hora vacía para primera ejecución

        # Crear hoja Historial vacía
        wb.create_sheet("Historial")

        wb.save(excel_path)
        log_message(f"Archivo Excel creado en {excel_path}")

    return excel_path
